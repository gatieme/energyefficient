#!/usr/bin/python
# encoding=utf-8


#!coding:utf-8

import re
import sys
import argparse
import commands
import os
import subprocess
import parse
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np


import openpyxl

class PerfPlotData :
    plotName = "none name"
    logfile = "none path"
    xData = []
    yData = []

    def __init__(self, plotName, logFile, xData, yData, color, marker) :

        """
        namelist        plot数据的标识, 对应各个调度算法[bl-switch, iks, hmp, hmpdb]
        xdata           存储了横轴的数据
        """
        self.plotName = plotName
        self.logFile  = logFile
        self.xData = xData
        self.yData = yData
        self.color = color
        self.marker = marker


def ShowPerfPlot(plotDataList, poly):
    #http://blog.csdn.net/kkxgx/article/details/6951959
    #http://www.mamicode.com/info-detail-280610.html
    #http://blog.csdn.net/panda1234lee/article/details/52311593
    #  中文信息支持
    mpl.rcParams['font.sans-serif'] = ['SimHei'] #用来正常显示中文标签
    mpl.rcParams['axes.unicode_minus'] = False #用来正常显示负号
    #自动调整label显示方式，如果太挤则倾斜显示
    fig = plt.figure(num = 1, figsize = (6, 5))
    fig.autofmt_xdate( )
    #plt.title("Scheduler Bench Performance...")
    title_str = "scheduler performance (perf/messaging) GROUP=100"
    plt.title(title_str)
    plt.xlabel("loop", size = 14)
    plt.ylabel("time", size = 14)
    plt.grid( ) # 开启网格

    for data in plotDataList :
        #  设置图表的信息
        print len(data.xData), len(data.yData)
        #  曲线平滑--http://rys1314520.iteye.com/blog/1820777
        #  曲线平滑--http://blog.sina.com.cn/s/blog_142e602960102wegx.html
        if poly == True :
            #计算多项式
            c = np.polyfit(data.xData, data.yData, 10)   #  拟合多项式的系数存储在数组c中
            yy = np.polyval(c, data.xData)                  #  根据多项式求函数值

            #进行曲线绘制
            x_new = np.linspace(0, 1000000, 11)
            f_liner = np.polyval(c,x_new)
            plt.plot(x_new, f_liner, color = data.color, linestyle = '--', marker = data.marker, label = data.plotName)
        else :
            plt.plot(data.xData, data.yData, color = data.color, linestyle = '--', marker = data.marker, label = data.plotName)
        plt.legend(loc = "upper left")

    #plt.savefig(title_str + '.png', format = 'png')
    plt.show( )



def ParsePlotData(str) :
    # 测试字符串格式化
    # 通过parse库可以实现与format相反的功能
    # 其结果类似与C语言的sscanf
    str_format =  "{:s}{:d}, {:f}"
    xydata = parse.parse(str_format, str)
    #print xydata
    return xydata #(xydata[1], xydata[2])



def ReadPlotXData(minData, maxData, step) :
    #  生成X轴的数据，从minData~maxData，步长为step
    xData = range(minData, maxData,  step)
    return xData;


def ReadPlotData(filepath, lines, iszero) :
    fileobject = open(filepath)

    if iszero == True :
        xData = [ 0 ]
        yData = [ 0 ]
    else :
        xData = [ ]
        yData = [ ]
    while 1 :
        linedata = fileobject.readlines(lines)

        if not linedata:
            break
        for line in linedata:
            #print line
            xyData = ParsePlotData(line)
            if (xyData != None) :
                #print "data = ", xyData[0], xyData[1]
                xData.append(xyData[1])
                yData.append(xyData[2])
            else :
                #print "line = ", line
                pass
    return (xData, yData)


#def PerfBenchPlotRun(nameTuple, colorTuple, marketTuple, bench, ming, maxg, setg) :
def PerfBenchPlotRun(nameTuple, colorTuple, marketTuple, args) :
    plotDataList = []
    #for name in nameTuple :
    for index in range(len(nameTuple)) :
        name = nameTuple[index]
        color = colorTuple[index]
        marker = markerTuple[index]
        if (name == "NULL") :
            break
        resultfile = args.directory + "/" + name + "/perf/" + args.bench + "/" \
                   + args.min_group + "-" + args.max_group + "-" + args.step_group + "-" +args.loop + ".log"
        print "\n=========================================="
        print "resultfile :", resultfile

        if ((int(args.min_group) + int(args.step_group)) > int(args.max_group)) :  #  同一个循环多次
            iszero = False
        else :
            iszero = True
        (xData, yData) = ReadPlotData(resultfile, 1000, iszero)
        print "+++++++", len(xData), len(yData), "+++++++"
        print xData
        print yData
        print "==========================================\n"
        plotdata = PerfPlotData(name, resultfile, xData, yData, color, marker)
        plotDataList.append(plotdata)

    #ShowPerfPlot(plotDataList, False)

    filename = args.bench + ".xlsx";
    sheetname = args.min_group + "-" + args.max_group + "-" + args.step_group + "-" + args.loop
    WriteExcelFile(plotDataList, filename, sheetname)


def WriteExcelFile(plotDataList, filename, sheetname) :
    """
    将数据写入Excel
    """
    #wb = openpyxl.Workbook()
    #ws = wb.active
    wb = openpyxl.load_workbook(filename)
    print wb.get_sheet_names( )
    if sheetname in wb.get_sheet_names( ) :
        ws = wb.get_sheet_by_name(sheetname)
        print sheetname, "in", wb.get_sheet_names( ) 
    else :
        ws = wb.create_sheet(sheetname)
        print "create", sheetname, "sheet"

    for row in range(len(plotDataList)) :
        ydata = plotDataList[row].yData
        ws.cell(row = row + 1, column = 1).value = str.upper(plotDataList[row].plotName)

        for col in range(len(plotDataList[row].yData)) :
            if row == 0:
                ws.cell(row = 1, column = col + 2).value = plotDataList[row].xData[col]
            else :
                ws.cell(row = row + 1, column = col + 2).value = ydata[col]
    wb.save(filename)



if __name__ == "__main__" :

#python logplot.py -d ../bench  -b messaging -min 10 -max 100 -step 10 -l 5
    reload(sys)
    sys.setdefaultencoding("utf-8")

    if len(sys.argv) > 1:               #  如果在程序运行时，传递了命令行参数
        pass
        #  打印传递的命令行参数的信息
        #print "您输入的所有参数共 %d 个，信息为 sys.argv = %s" % (len(sys.argv), sys.argv)

        #for i, eachArg in enumerate(sys.argv):
        #    print "[%d] = %s" % (i, eachArg)
    else:
        print "Useage : read.py file..."
        exit(0)

    parser = argparse.ArgumentParser( )
    #parser.add_argument("-n", "--name", dest = "name", help = "bl-switch | iks | hmp | hmpdb...")
    parser.add_argument("-b", "--bench", dest = "bench", help = "messaging | pipe...")
    parser.add_argument("-d", "--dir", dest = "directory", help = "The Directory")
    parser.add_argument("-f", "--file", dest = "resultfile", help = "The file you want to read...")
    parser.add_argument("-min", "--min_group", dest = "min_group", help = "The min group you give...")
    parser.add_argument("-max", "--max_group", dest = "max_group", help = "The max group you give...")
    parser.add_argument("-step", "--step_group", dest = "step_group", help = "The step of the group grown you  give...")
    parser.add_argument("-l", "--loop", dest = "loop", help = "The file you want to read...")
    args = parser.parse_args( )

    #nameTuple = ( "hmp", "hmpdb")
    nameTuple = ( "bl-switch", "iks", "hmp", "hmpdb")
    #nameTuple = ( "bl-switch", "iks", "hmp", "hmpdb", "little-cluster", "big-cluster", "big-little-cluster")
    #nameTuple = ( "little-cluster", "big-cluster", "big-little-cluster")
    #   1）控制颜色
    #   颜色之间的对应关系为
    #   b---blue   c---cyan  g---green    k----black
    #   m---magenta r---red  w---white    y----yellow
    colorTuple = ( 'b', 'c', 'g', 'k', 'm', 'r', 'y', 'y')
    #.  Point marker
    #,  Pixel marker
    #o  Circle marker
    #v  Triangle down marker
    #^  Triangle up marker
    #<  Triangle left marker
    #>  Triangle right marker
    #1  Tripod down marker
    #2  Tripod up marker
    #3  Tripod left marker
    #4  Tripod right marker
    #s  Square marker
    #p  Pentagon marker
    #*  Star marker
    #h  Hexagon marker
    #H  Rotated hexagon D Diamond marker
    #d  Thin diamond marker
    #| Vertical line (vlinesymbol) marker
    #_  Horizontal line (hline symbol) marker
    #+  Plus marker
    #x  Cross (x) marker
    markerTuple= ( 'o', '^', '*', 's', 'p', '2', 'h', )
    PerfBenchPlotRun(nameTuple, colorTuple, markerTuple, args)

    exit(0)
